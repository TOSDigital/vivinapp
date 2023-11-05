from django.shortcuts import render, reverse
from .models import Project, User, UserProfile, OfficeLogin, SiteEngineer, AttendanceRecord, LaborTypes, Contractor, MaterialCategory
from django.views.generic import ListView, UpdateView, DeleteView, DetailView, TemplateView, CreateView, View, FormView
from .forms import *
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Sum, F, DecimalField, Q
from decimal import Decimal

# Create your views here.

# Proejct Views
class DashboardView(TemplateView):
    template_name = 'dashboard.html'

class ProjectListView(ListView):
    template_name = 'projects.html'
    queryset = Project.objects.all().order_by('-id')
    context_object_name = 'project'

class ProjectCreateView(CreateView):
    template_name = 'projectcreate.html'
    form_class = ProjectCreationForm

    def get_success_url(self):
        return reverse("projects")

class ProjectUpdateView(UpdateView):
    template_name = "projectupdate.html"
    model = Project
    form_class = ProjectCreationForm
    context_object_name = 'project'

    
    def get_success_url(self):
        return reverse("projects")

class ProjectDeatilView(DetailView):
    template_name = "projectdetail.html"
    model = Project
    context_object_name = 'projectdetail'

# project Views End 

# Attandance Feature start
class AttendanceRecordCreateView(CreateView):
    model = AttendanceRecord
    form_class = AttendanceRecordForm
    template_name = 'attendance_record.html'
    
    def get_success_url(self):
        return reverse("attendance_recordlist")

def load_contractors(request):
    project_id = request.GET.get('project')
    if project_id:  # Check if project_id is not None or empty string
        contractors = Contractor.objects.filter(Project_id=project_id).order_by('Contractor_name')
    else:
        contractors = Contractor.objects.none()
    return render(request, 'partials/contractor_dropdown_list_options.html', {'contractors': contractors})

def load_labor_types(request):
    contractor_id = request.GET.get('contractor')
    labor_types = LaborTypes.objects.filter(Contractor_id=contractor_id).order_by('Labor_type')
    return render(request, 'partials/labor_type_dropdown_list_options.html', {'labor_types': labor_types})

class AttendanceRecordListView(ListView):
    model = AttendanceRecord
    queryset = AttendanceRecord.objects.all().order_by('-id')
    template_name = "attendance_record_list.html"
    context_object_name = 'recordlist'

class AttendanceRecordUpdateView(UpdateView):
    template_name = "attendance_record_list_update.html"
    model = AttendanceRecord
    context_object_name = 'attendance'
    fields = (
        'number_of_workers',
        'Remarks'
    )

    def get_success_url(self):
        return reverse("attendance_recordlist")

class AttendanceRecordDeleteView(DeleteView):
    model = AttendanceRecord
    context_object_name = 'attendancerecord'
    template_name = "attendance_record_delete.html"

    def get_success_url(self):
        return reverse("attendance_recordlist")

# Attandance Views End 

# Reports start

class AllReports(TemplateView):
    template_name = "reports.html"

# Reports end

# Attandance report View 
class WageCalculationView(FormView):
    form_class = WageCalculationForm
    template_name = 'laborreport.html'

    def form_valid(self, form):
        project = form.cleaned_data['project']
        contractor = form.cleaned_data['contractor']
        from_date = form.cleaned_data['from_date']
        to_date = form.cleaned_data['to_date']

        # Fetch labor types for the contractor
        labor_types = LaborTypes.objects.filter(Contractor=contractor)
        # ...
        labor_types_with_wages = []

        for labor_type in labor_types:
            attendance_records = AttendanceRecord.objects.filter(
                project=project,
                contractor=contractor,
                labor_type=labor_type,
                date__range=(from_date, to_date)
            )
            total_workers = attendance_records.aggregate(
                total_workers=Sum('number_of_workers')
            )['total_workers'] or 0
            row_total = labor_type.wage * total_workers
            labor_types_with_wages.append({
                'labor_type': labor_type,
                'wage': labor_type.wage,
                'total_workers': total_workers,
                'row_total': row_total
            })

        total_wage = sum(ltw['row_total'] for ltw in labor_types_with_wages)

        # Check if the request is to download the excel
        if 'download_excel' in self.request.POST:
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename={date}-wage-report.xlsx'.format(
                date=timezone.now().strftime('%Y-%m-%d'),
            )
            wb = Workbook()
            ws = wb.active
            ws.title = 'Wage Report'

            # Write table headers
            columns = ['Labor Type', 'Wage per Worker', 'Number of Workers', 'Row Total']
            for col_num, column_title in enumerate(columns, start=1):
                ws.cell(row=1, column=col_num).value = column_title

            # Write table rows
            for row_num, ltw in enumerate(labor_types_with_wages, start=2):
                ws.cell(row=row_num, column=1).value = ltw['labor_type'].Labor_type
                ws.cell(row=row_num, column=2).value = ltw['wage']
                ws.cell(row=row_num, column=3).value = ltw['total_workers']
                ws.cell(row=row_num, column=4).value = ltw['row_total']

            # Write the total wage
            total_row_num = row_num + 1
            ws.cell(row=total_row_num, column=3).value = 'Total'
            ws.cell(row=total_row_num, column=4).value = total_wage

            # Add project and contractor details after the table
            details_start_row = total_row_num + 2
            ws.cell(row=details_start_row, column=1).value = 'Project Name:'
            ws.cell(row=details_start_row, column=2).value = project.Project_name

            ws.cell(row=details_start_row + 1, column=1).value = 'Contractor Name:'
            ws.cell(row=details_start_row + 1, column=2).value = contractor.Contractor_name

            ws.cell(row=details_start_row + 2, column=1).value = 'From Date:'
            ws.cell(row=details_start_row + 2, column=2).value = str(from_date)

            ws.cell(row=details_start_row + 3, column=1).value = 'To Date:'
            ws.cell(row=details_start_row + 3, column=2).value = str(to_date)

            wb.save(response)
            return response

        context = self.get_context_data(
            form=form,
            labor_types_with_wages=labor_types_with_wages,
            total_wage=total_wage,
            project=project,
            contractor=contractor,
            from_date=from_date,
            to_date=to_date
        )
        return self.render_to_response(context)

# Add this view to handle the AJAX request for the contractor dropdown
class ContractorOptionsView(View):
    def get(self, request, *args, **kwargs):
        project_id = request.GET.get('project')
        contractors = Contractor.objects.filter(Project_id=project_id)
        return JsonResponse(list(contractors.values('id', 'Contractor_name')), safe=False)

class OvertimeWageCalculationView(FormView):
    form_class = OvertimeWageCalculationForm
    template_name = 'overtime_wage_calculation.html'

    def form_valid(self, form):
        # Extract form data
        project = form.cleaned_data['project']
        contractor = form.cleaned_data['contractor']
        from_date = form.cleaned_data['from_date']
        to_date = form.cleaned_data['to_date']

        # Query labor types and related overtime records
        labor_types = LaborTypes.objects.filter(Contractor=contractor)
        overtime_data = []
        total_wage = Decimal('0.00')

        # Calculate overtime wages
        for labor_type in labor_types:
            overtime_records = OvertimeRecord.objects.filter(
                labor_type=labor_type,
                date__range=(from_date, to_date)
            )
            total_hours = overtime_records.aggregate(total_hours=Sum('number_of_hours'))['total_hours'] or 0
            total_workers = overtime_records.aggregate(total_workers=Sum('number_of_workers'))['total_workers'] or 0
            total_overtime_wage = total_hours * total_workers * labor_type.overtime_wage
            overtime_data.append({
                'labor_type': labor_type.Labor_type,
                'overtime_wage': labor_type.overtime_wage,
                'total_hours': total_hours,
                'total_workers': total_workers,
                'total_overtime_wage': total_overtime_wage
            })
            total_wage += total_overtime_wage

        # Check if the user has requested the Excel file download
        if 'download_excel' in self.request.POST:
            return self.handle_excel_download(project, contractor, from_date, to_date, overtime_data, total_wage)
        
        # Render the results in the template if not downloading
        context = self.get_context_data(
            form=form,
            overtime_data=overtime_data,
            total_wage=total_wage,
            project=project,
            contractor=contractor,
            from_date=from_date,
            to_date=to_date
        )
        return self.render_to_response(context)

    def handle_excel_download(self, project, contractor, from_date, to_date, overtime_data, total_wage):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{project.Project_name}_{from_date}_{to_date}_overtime_wages.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write table headers
        columns = ['Labor Type', 'Overtime Wage per Hour', 'Total Hours', 'Total Workers', 'Total Overtime Wage']
        for col_num, column_title in enumerate(columns, start=1):
            ws.cell(row=1, column=col_num).value = column_title

        # Write table rows
        for row_num, item in enumerate(overtime_data, start=2):
            ws.cell(row=row_num, column=1).value = item['labor_type']
            ws.cell(row=row_num, column=2).value = item['overtime_wage']
            ws.cell(row=row_num, column=3).value = item['total_hours']
            ws.cell(row=row_num, column=4).value = item['total_workers']
            ws.cell(row=row_num, column=5).value = item['total_overtime_wage']

        # Add a row for the total wage
        total_row_num = row_num + 1
        ws.cell(row=total_row_num, column=4).value = 'Total Overtime Wage'
        ws.cell(row=total_row_num, column=5).value = total_wage

        # Add the project, contractor, and date details after the table
        details_start_row = total_row_num + 2
        ws.cell(row=details_start_row, column=1).value = f'Project: {project.Project_name}'
        ws.cell(row=details_start_row + 1, column=1).value = f'Contractor: {contractor.Contractor_name}'
        ws.cell(row=details_start_row + 2, column=1).value = f'From Date: {from_date.strftime("%Y-%m-%d")}'
        ws.cell(row=details_start_row + 3, column=1).value = f'To Date: {to_date.strftime("%Y-%m-%d")}'

        # Auto-adjust columns' width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        wb.save(response)
        return response


class LoadContractorsView(View):
    def get(self, request, *args, **kwargs):
        project_id = request.GET.get('project_id')
        contractors = Contractor.objects.filter(Project_id=project_id).order_by('Contractor_name')
        return JsonResponse(list(contractors.values('id', 'Contractor_name')), safe=False)

# Attandance report View end

# contractor views start

class ContractorListView(ListView):
    template_name = "contractorlist.html"
    model = Contractor
    queryset = Contractor.objects.all().order_by('-id')
    context_object_name = "contractors"


class ContractorCreateView(CreateView):
    template_name = "contractorcreate.html"
    model = Contractor
    form_class = ContractorForm

    def get_success_url(self):
        return reverse("attendance_recordlist")

class ContractorUpdateView(UpdateView):
    template_name = "contractorupdate.html"
    model = Contractor
    form_class = ContractorForm

    def get_success_url(self):
        return reverse("attendance_recordlist")

class ContractorDeleteView(DeleteView):
    template_name = "contractordelete.html"
    model = Contractor
    context_object_name = "contractor"

    def get_success_url(self):
        return reverse("attendance_recordlist")

# contractor views end

# Labortype Views start


class LaborTypeListView(ListView):
    template_name = "labortypelist.html"
    model = LaborTypes
    queryset = LaborTypes.objects.all().order_by('-id')
    context_object_name = "labortypes"

class LaborTypeCreateView(CreateView):
    template_name = "labortypecreate.html"
    model = LaborTypes
    form_class = LaborTypeForm

    def save(self, *args, **kwargs):
        if self.overtime_wage is None or self.overtime_wage == '':
            self.overtime_wage = self.calculate_overtime_wage()
        super(LaborTypes, self).save(*args, **kwargs)

    def get_success_url(self):
        return reverse("labortypelist")

class LaborTypeUpdateView(UpdateView):
    template_name = "labortypeupdate.html"
    model = LaborTypes
    form_class = LaborTypeForm

    def get_success_url(self):
        return reverse("labortypelist")

class LaborTypeDeleteView(DeleteView):
    template_name = "labortypedelete.html"
    model = LaborTypes
    context_object_name = "labortype"

    def get_success_url(self):
        return reverse("labortypelist")

class LaborTypeDeatilView(DetailView):
    template_name = "labortypeview.html"
    model = LaborTypes
    context_object_name = "labor"

# Labortype Views end

# overtime Views start

class OvertimeListView(ListView):
    model = OvertimeRecord
    template_name = "overtimelist.html"
    context_object_name = "overtimewage"

class OvertimeCreateView(CreateView):
    model = OvertimeRecord
    template_name = "overtimecreate.html"
    form_class = OvertimeRecordForm

    def get_success_url(self):
        return reverse("overtimelist")

def load_contractors(request):
    project_id = request.GET.get('project')
    if project_id:  # Check if project_id is not None or empty string
        contractors = Contractor.objects.filter(Project_id=project_id).order_by('Contractor_name')
    else:
        contractors = Contractor.objects.none()
    return render(request, 'partials/contractor_dropdown_list_options.html', {'contractors': contractors})

def load_labor_types(request):
    contractor_id = request.GET.get('contractor')
    labor_types = LaborTypes.objects.filter(Contractor_id=contractor_id).order_by('Labor_type')
    return render(request, 'partials/labor_type_dropdown_list_options.html', {'labor_types': labor_types})

class OvertimeUpdateView(UpdateView):
    model = OvertimeRecord
    template_name = "overtimeupdate.html"
    fields = (
        'number_of_workers', 'number_of_hours', 'Remarks'
    )
    context_object_name = "overtime"

    def get_success_url(self):
        return reverse("overtimelist")

class OvertimeDeleteView(DeleteView):
    model = OvertimeRecord
    template_name = "overtimedelete.html"
    context_object_name = "overtime"

    def get_success_url(self):
        return reverse("overtimelist")

# overtime view end

# Indents View
class MaterialCategoryList(ListView):
    template_name = "materialcategorylist.html"
    queryset = MaterialCategory.objects.all()
    context_object_name = "materialcategories"
    
class MaterialCreateView(CreateView):
    model = MaterialCategory
    template_name = "materialcategorycreate.html"
    form_class = MaterialCategoryForm

    def get_success_url(self):
        return reverse("materialcategorylist")

class MaterialDeleteView(DeleteView):
    model = MaterialCategory
    template_name = "materialcategorydelete.html"
    context_object_name = "materialcategory"

    def get_success_url(self):
        return reverse("materialcategorylist")


    









 
    



